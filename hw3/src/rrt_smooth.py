import numpy as np
import matplotlib.pyplot as plt
import cv2
import random
import openpyxl
import open3d as o3d

class Nodes:
    """Class to store the RRT graph"""
    def __init__(self, x,y):
        self.x = x
        self.y = y
        self.parent_x = []
        self.parent_y = []

start_point = []
def click_event(event, x, y, flags, params):

    global start_point
    if event == cv2.EVENT_LBUTTONDBLCLK:
        start_point = [x, y]
        cv2.circle(map, start_point, 3, (0, 0, 255), -1)
        cv2.imshow('map', map)

def find_nearest_point(node_list, target):
    temp_dist = []
    for j in range(len(node_list)):
        dist = np.linalg.norm(np.array(target) - np.array([node_list[j].x, node_list[j].y]))
        temp_dist.append(dist)
    near_idx = temp_dist.index(min(temp_dist))
    nearest = [node_list[near_idx].x, node_list[near_idx].y]
    return near_idx, nearest

def add_new_node(node_list, i, new_point, near_idx):
    node_list.append(i)
    node_list[i] = Nodes(new_point[0],new_point[1])
    node_list[i].parent_x = node_list[near_idx].parent_x.copy()
    node_list[i].parent_y = node_list[near_idx].parent_y.copy()
    node_list[i].parent_x.append(new_point[0])
    node_list[i].parent_y.append(new_point[1])
    i = i + 1
    cv2.circle(map, new_point, 3, (0, 255, 0), -1)
    cv2.line(map, nearest, new_point, (0, 255, 0), 1)
    return node_list, i

def generate_point(target, nearest):
    vecter = np.array(target)-np.array(nearest)
    if target == nearest:
        new_point = nearest
    else:
        length = np.linalg.norm(vecter)
        vecter = (vecter / length) * min(step_size, length)
        new_point = [int(nearest[0]+vecter[0]), int(nearest[1]+vecter[1])]
    return new_point

if __name__ == "__main__":

    map = cv2.imread("map.png")
    map_bw = cv2.imread("map.png",0)

    #####image erosion to make obstacle more transparent#####
    kernel = np.ones((5,5), np.uint8)
    erosion = cv2.erode(map_bw, kernel, iterations = 2)

    #####get object color rgb#####
    object = input("I want to search: ")
    wb = openpyxl.load_workbook('category.xlsx', data_only=True)
    s1 = wb['Sheet1']
    for i in range(2,103):
        if s1.cell(i,5).value == object:
            object_color = s1.cell(i,2).value
            break
    wb.save('category.xlsx')
    object_color = object_color.replace("(","").replace(")","").replace(",","")
    object_color = np.array([int(temp)for temp in object_color.split() if temp.isdigit()]) # take the numbers
    # print(object_color)

    #####get the mean position pixel of the object#####
    object_point = []
    for i in range(map.shape[0]):
        for j in range(map.shape[1]):
            if map[i,j,0] == object_color[2] and map[i,j,1] == object_color[1] and map[i,j,2] == object_color[0]:
                object_point.append([i,j])
    # dst = map[np.where((map[:,:,0] == object_color[2]) & (map[:,:,1] == object_color[1]) & (map[:,:,2] == object_color[0]))]
    object_mean = np.round(np.mean(np.array(object_point), axis=0)).astype(int)
    # print("mean point: ", object_mean)

    #####choose a start point#####
    cv2.imshow("map",map)
    cv2.setMouseCallback('map', click_event)
    cv2.waitKey()
    cv2.destroyAllWindows()

    #####choose the end point(hard code)#####
    if object == "refrigerator":
        end_point = [object_mean[1]+2, object_mean[0]+15]
        cv2.circle(map, end_point, 3, (0, 0, 0), -1)
    elif object == "rack":
        end_point = [object_mean[1]-16, object_mean[0]+4]
        cv2.circle(map, end_point, 3, (0, 0, 0), -1)
    elif object == "cushion":
        end_point = [object_mean[1]-27, object_mean[0]-26]
        cv2.circle(map, end_point, 3, (0, 0, 0), -1)
    elif object == "lamp":
        end_point = [object_mean[1]-23, object_mean[0]-1]
        cv2.circle(map, end_point, 3, (0, 0, 0), -1)
    elif object == "cooktop":
        end_point = [object_mean[1]+1, object_mean[0]-28]
        cv2.circle(map, end_point, 3, (0, 0, 0), -1)




    # initialize
    step_size = 15
    node_list = [0]
    
    # insert the starting point in the node class
    node_list[0] = Nodes(start_point[0],start_point[1])
    node_list[0].parent_x.append(start_point[0])
    node_list[0].parent_y.append(start_point[1])
    cv2.circle(map, start_point, 3, (0, 0, 255), -1)

    ##### RRT algorithm #####
    i = 1
    find_path = 0
    while True:
        Xrand = [random.randint(0,map.shape[1]), random.randint(68,map.shape[0])] # set a random point

        # find closest point
        near_idx, nearest = find_nearest_point(node_list, Xrand)

        # generate new point
        new_point = generate_point(Xrand, nearest)
        
        if erosion[new_point[1]-1, new_point[0]-1] == 255:
            # save nodes
            node_list, i = add_new_node(node_list, i, new_point, near_idx)
            
            # find nearest point to end point
            near_idx, nearest = find_nearest_point(node_list, end_point)
            new_point = generate_point(end_point, nearest)
            while erosion[new_point[1]-1, new_point[0]-1] == 255:
                node_list, i = add_new_node(node_list, i, new_point, near_idx)
                new_point = generate_point(end_point, new_point)

                if new_point == end_point:
                    node_list, i = add_new_node(node_list, i, new_point, near_idx)
                    i = i - 1
                    path_i = i
                    print("new point: ", new_point)
                    print("Find path")
                    for j in range(len(node_list[i].parent_x)-1):
                        cv2.line(map, (int(node_list[i].parent_x[j]),int(node_list[i].parent_y[j])), (int(node_list[i].parent_x[j+1]),int(node_list[i].parent_y[j+1])), (0,0,255), 2)
                    find_path = 1
                    break
        
        cv2.imshow("map",map)
        cv2.waitKey(1)

        if find_path:
            break
    
    cv2.imshow("map",map)
    cv2.waitKey()
    cv2.destroyAllWindows()


    # point = np.load('semantic_3d_pointcloud/point.npy')
    # color = np.load('semantic_3d_pointcloud/color01.npy')

    # # remove *(15/496)the ceiling and the floor
    # point_cut = point[np.where((point[:,1]>(-0.03)) & (point[:,1]<(-0.005)))]
    # color_cut = color[np.where((point[:,1]>(-0.03)) & (point[:,1]<(-0.005)))]
    # point_cut = point_cut*10000./255.

    # for i in range(len(node_list[path_i].parent_x)):
    #     pixel2real_x = int(node_list[path_i].parent_x[i])*(15/496)
    #     pixel2real_y = int(node_list[path_i].parent_y[i])*(11/369)
    #     row = np.array([(-pixel2real_y)+7, -0.5, pixel2real_x-5]) # translation hard code
    #     point_cut = np.append(point_cut, [row],axis= 0)
    #     color_cut = np.append(color_cut, [[0,1,0]],axis= 0)

   
    # pcd = o3d.geometry.PointCloud()
    # pcd.points = o3d.utility.Vector3dVector(point_cut)
    # pcd.colors = o3d.utility.Vector3dVector(color_cut)
    # o3d.visualization.draw_geometries([pcd])