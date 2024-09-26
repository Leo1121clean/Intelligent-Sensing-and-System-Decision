import numpy as np
from PIL import Image
import numpy as np
import habitat_sim
import cv2
import json
import random
import openpyxl
from scipy.spatial.transform import Rotation as R
import math
from scipy.io import loadmat

class Nodes:
    def __init__(self, x,y):
        self.x = x
        self.y = y
        self.parent_x = []
        self.parent_y = []

start_point = []
def click_event(event, x, y, flags, params):
    global start_point
    if event == cv2.EVENT_LBUTTONDBLCLK:
        start_point = [x, y]
        cv2.circle(map, start_point, 3, (0, 0, 255), -1)
        cv2.imshow('map', map)

def find_nearest_point(node_list, target):
    temp_dist = []
    for j in range(len(node_list)):
        dist = np.linalg.norm(np.array(target) - np.array([node_list[j].x, node_list[j].y]))
        temp_dist.append(dist)
    near_idx = temp_dist.index(min(temp_dist))
    nearest = [node_list[near_idx].x, node_list[near_idx].y]
    return near_idx, nearest

def add_new_node(node_list, i, new_point, near_idx):
    node_list.append(i)
    node_list[i] = Nodes(new_point[0],new_point[1])
    node_list[i].parent_x = node_list[near_idx].parent_x.copy()
    node_list[i].parent_y = node_list[near_idx].parent_y.copy()
    node_list[i].parent_x.append(new_point[0])
    node_list[i].parent_y.append(new_point[1])
    i = i + 1
    cv2.circle(map, new_point, 3, (0, 255, 0), -1)
    cv2.line(map, nearest, new_point, (0, 255, 0), 1)
    return node_list, i

def generate_point(target, nearest):
    vecter = np.array(target)-np.array(nearest)
    if target == nearest:
        new_point = nearest
    else:
        length = np.linalg.norm(vecter)
        vecter = (vecter / length) * min(step_size, length)
        new_point = [int(nearest[0]+vecter[0]), int(nearest[1]+vecter[1])]
    return new_point

#####image imput#####
map = cv2.imread("map.png")
map_bw = cv2.imread("map.png",0)

#####image erosion to make obstacle more transparent#####
kernel = np.ones((7,7), np.uint8)
erosion = cv2.erode(map_bw, kernel, iterations = 2)

#####get object color rgb#####
object = input("I want to search: ")
wb = openpyxl.load_workbook('category.xlsx', data_only=True)
s1 = wb['Sheet1']
for i in range(2,103):
    if s1.cell(i,5).value == object:
        object_color = s1.cell(i,2).value
        break
wb.save('category.xlsx')
object_color = object_color.replace("(","").replace(")","").replace(",","")
object_color = np.array([int(temp)for temp in object_color.split() if temp.isdigit()]) # take the numbers

#####get the mean position pixel of the object#####
object_point = []
for i in range(map.shape[0]):
    for j in range(map.shape[1]):
        if map[i,j,0] == object_color[2] and map[i,j,1] == object_color[1] and map[i,j,2] == object_color[0]:
            object_point.append([i,j])
object_mean = np.round(np.mean(np.array(object_point), axis=0)).astype(int)
# print("mean point: ", object_mean)

#####choose a start point#####
cv2.imshow("map",map)
cv2.setMouseCallback('map', click_event)
cv2.waitKey()
cv2.destroyAllWindows()

#####choose the end point(hard code)#####
if object == "refrigerator":
    end_point = [object_mean[1]+2, object_mean[0]+15]
    cv2.circle(map, end_point, 3, (0, 0, 0), -1)
elif object == "rack":
    end_point = [object_mean[1]-16, object_mean[0]+4]
    cv2.circle(map, end_point, 3, (0, 0, 0), -1)
elif object == "cushion":
    end_point = [object_mean[1]-27, object_mean[0]-26]
    cv2.circle(map, end_point, 3, (0, 0, 0), -1)
elif object == "lamp":
    end_point = [object_mean[1]-23, object_mean[0]-1]
    cv2.circle(map, end_point, 3, (0, 0, 0), -1)
elif object == "cooktop":
    end_point = [object_mean[1]+1, object_mean[0]-28]
    cv2.circle(map, end_point, 3, (0, 0, 0), -1)


# initialize
step_size = 15
node_list = [0]
node_list_b = [0]
record_path = []

# insert the starting point in the node class
node_list[0] = Nodes(start_point[0],start_point[1])
node_list[0].parent_x.append(start_point[0])
node_list[0].parent_y.append(start_point[1])
cv2.circle(map, start_point, 3, (0, 0, 255), -1)

# insert the end point in the node class b
node_list_b[0] = Nodes(end_point[0],end_point[1])
node_list_b[0].parent_x.append(end_point[0])
node_list_b[0].parent_y.append(end_point[1])
cv2.circle(map, end_point, 3, (0, 0, 255), -1)

##### RRT algorithm #####
i = 1
ib = 1
find_path = 0
while True:
    Xrand = [random.randint(0,map.shape[1]), random.randint(68,map.shape[0])] # set a random point

    # find closest point
    near_idx, nearest = find_nearest_point(node_list, Xrand)

    # generate new point
    new_point = generate_point(Xrand, nearest)
    
    if erosion[new_point[1]-1, new_point[0]-1] == 255:
        # save nodes
        node_list, i = add_new_node(node_list, i, new_point, near_idx)
        
        # find nearest point to target, and extend
        target = new_point
        near_idx, nearest = find_nearest_point(node_list_b, target)
        new_point = generate_point(target, nearest)
        while erosion[new_point[1]-1, new_point[0]-1] == 255:
            node_list_b, ib = add_new_node(node_list_b, ib, new_point, near_idx)
            new_point = generate_point(target, new_point)

            cv2.imshow("map",map)
            cv2.waitKey(1)

            dist = np.linalg.norm(np.array(target) - np.array(new_point))
            if dist < step_size:
                print("Match the tree")
                ib = ib - 1
                i = i - 1
                for j in range(len(node_list_b[ib].parent_x)-1):
                    record_path.append([node_list_b[ib].parent_x[j], node_list_b[ib].parent_y[j]])
                    # cv2.line(map, (int(node_list_b[ib].parent_x[j]),int(node_list_b[ib].parent_y[j])), (int(node_list_b[ib].parent_x[j+1]),int(node_list_b[ib].parent_y[j+1])), (0,0,255), 2)
                for j in range(len(node_list[i].parent_x)-1):
                    record_path.append([node_list[i].parent_x[len(node_list[i].parent_x)-1-j], node_list[i].parent_y[len(node_list[i].parent_x)-1-j]])
                    # cv2.line(map, (int(node_list[i].parent_x[j]),int(node_list[i].parent_y[j])), (int(node_list[i].parent_x[j+1]),int(node_list[i].parent_y[j+1])), (255,0,0), 2)
                for j in range(len(record_path)-1):
                    cv2.line(map, record_path[j], record_path[j+1], (0,0,255), 2)
                if record_path[0] == end_point:
                    record_path.reverse() # make the end of path list is the end point
                find_path = 1
                break        

    node_list, node_list_b = node_list_b, node_list
    i, ib = ib, i
    
    cv2.imshow("map",map)
    cv2.waitKey(1)

    if find_path:
        break

cv2.imshow("map",map)
cv2.waitKey()
cv2.destroyAllWindows()

# save navigation point
navigation_point = []
for i in range(len(record_path)):
    pixel2real_x = int(record_path[i][0])*(15/map.shape[1])
    pixel2real_y = int(record_path[i][1])*(11/map.shape[0])
    navigation_point.append([(-pixel2real_y)+7, pixel2real_x-5]) # translation hard code

################################################################################################################################

# load colors
colors = loadmat('color101.mat')['colors']
colors = np.insert(colors, 0, values=np.array([[0,0,0]]), axis=0) #to make the color be correct

# This is the scene we are going to load.
# support a variety of mesh formats, such as .glb, .gltf, .obj, .ply
### put your scene path ###
test_scene = "apartment_0/habitat/mesh_semantic.ply"
path = "apartment_0/habitat/info_semantic.json"

#global test_pic
#### instance id to semantic id 
with open(path, "r") as f:
    annotations = json.load(f)

id_to_label = []
instance_id_to_semantic_label_id = np.array(annotations["id_to_label"])
for i in instance_id_to_semantic_label_id:
    if i < 0:
        id_to_label.append(0)
    else:
        id_to_label.append(i)
id_to_label = np.asarray(id_to_label)



######

sim_settings = {
    "scene": test_scene,  # Scene path
    "default_agent": 0,  # Index of the default agent
    "sensor_height": 1.5,  # Height of sensors in meters, relative to the agent
    "width": 512,  # Spatial resolution of the observations
    "height": 512,
    "sensor_pitch": 0,  # sensor pitch (x rotation in rads)
}

# This function generates a config for the simulator.
# It contains two parts:
# one for the simulator backend
# one for the agent, where you can attach a bunch of sensors

def transform_rgb_bgr(image):
    return image[:, :, [2, 1, 0]]

def transform_depth(image):
    depth_img = (image / 10 * 255).astype(np.uint8)
    return depth_img

def transform_semantic(semantic_obs):
    semantic_img = Image.new("P", (semantic_obs.shape[1], semantic_obs.shape[0]))
    semantic_img.putpalette(colors.flatten())
    semantic_img.putdata(semantic_obs.flatten().astype(np.uint8))
    semantic_img = semantic_img.convert("RGB")
    semantic_img = cv2.cvtColor(np.asarray(semantic_img), cv2.COLOR_RGB2BGR)
    return semantic_img

def make_simple_cfg(settings):
    # simulator backend
    sim_cfg = habitat_sim.SimulatorConfiguration()
    sim_cfg.scene_id = settings["scene"]
    

    # In the 1st example, we attach only one sensor,
    # a RGB visual sensor, to the agent
    rgb_sensor_spec = habitat_sim.CameraSensorSpec()
    rgb_sensor_spec.uuid = "color_sensor"
    rgb_sensor_spec.sensor_type = habitat_sim.SensorType.COLOR
    rgb_sensor_spec.resolution = [settings["height"], settings["width"]]
    rgb_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    rgb_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    rgb_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    #depth snesor
    depth_sensor_spec = habitat_sim.CameraSensorSpec()
    depth_sensor_spec.uuid = "depth_sensor"
    depth_sensor_spec.sensor_type = habitat_sim.SensorType.DEPTH
    depth_sensor_spec.resolution = [settings["height"], settings["width"]]
    depth_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    depth_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    depth_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    #semantic snesor
    semantic_sensor_spec = habitat_sim.CameraSensorSpec()
    semantic_sensor_spec.uuid = "semantic_sensor"
    semantic_sensor_spec.sensor_type = habitat_sim.SensorType.SEMANTIC
    semantic_sensor_spec.resolution = [settings["height"], settings["width"]]
    semantic_sensor_spec.position = [0.0, settings["sensor_height"], 0.0]
    semantic_sensor_spec.orientation = [
        settings["sensor_pitch"],
        0.0,
        0.0,
    ]
    semantic_sensor_spec.sensor_subtype = habitat_sim.SensorSubType.PINHOLE

    # agent
    agent_cfg = habitat_sim.agent.AgentConfiguration()

    agent_cfg.sensor_specifications = [rgb_sensor_spec, depth_sensor_spec, semantic_sensor_spec]
    ##################################################################
    ### change the move_forward length or rotate angle
    ##################################################################
    agent_cfg.action_space = {
        "move_forward": habitat_sim.agent.ActionSpec(
            "move_forward", habitat_sim.agent.ActuationSpec(amount=0.05) # 0.01 means 0.01 m
        ),
        "turn_left": habitat_sim.agent.ActionSpec(
            "turn_left", habitat_sim.agent.ActuationSpec(amount=1.0) # 1.0 means 1 degree
        ),
        "turn_right": habitat_sim.agent.ActionSpec(
            "turn_right", habitat_sim.agent.ActuationSpec(amount=1.0)
        ),
    }

    return habitat_sim.Configuration(sim_cfg, [agent_cfg])


cfg = make_simple_cfg(sim_settings)
sim = habitat_sim.Simulator(cfg)


# initialize an agent
agent = sim.initialize_agent(sim_settings["default_agent"])

# Set agent state
agent_state = habitat_sim.AgentState()
# agent_state.position = np.array([0.0, 0.0, 0.0])  # agent in world space
agent_state.position = np.array([navigation_point[0][0], 0, navigation_point[0][1]])  # agent in start point
agent.set_state(agent_state)

# obtain the default, discrete actions that an agent can perform
# default action space contains 3 actions: move_forward, turn_left, and turn_right
action_names = list(cfg.agents[sim_settings["default_agent"]].action_space.keys())
print("Discrete action space: ", action_names)


FORWARD_KEY="w"
LEFT_KEY="a"
RIGHT_KEY="d"
FINISH="f"
print("#############################")
print("use keyboard to control the agent")
print(" w for go forward  ")
print(" a for turn left  ")
print(" d for trun right  ")
print(" f for finish and quit the program")
print("#############################")

# highlight the object
def object_highlight(rgb, sem):
    index = np.where((sem[:,:,2]==object_color[0]) & (sem[:,:,1]==object_color[1]) & (sem[:,:,0]==object_color[2]))
    if len(index[0]) != 0:
        rgb[index] = cv2.addWeighted(rgb[index], 0.6, sem[index], 0.4, 50)
    return rgb


def navigateAndSee(action=""):
    if action in action_names:
        observations = sim.step(action)
        #print("action: ", action)

        highlighted = object_highlight(transform_rgb_bgr(observations["color_sensor"]), transform_semantic(id_to_label[observations["semantic_sensor"]]))

        cv2.imshow("RGB", highlighted)
        # cv2.imshow("RGB", transform_rgb_bgr(observations["color_sensor"]))
        # cv2.imshow("depth", transform_depth(observations["depth_sensor"]))
        # cv2.imshow("semantic", transform_semantic(id_to_label[observations["semantic_sensor"]]))
        agent_state = agent.get_state()
        sensor_state = agent_state.sensor_states['color_sensor']
        # print("camera pose: x y z rw rx ry rz")
        # print(sensor_state.position[0],sensor_state.position[1],sensor_state.position[2],  sensor_state.rotation.w, sensor_state.rotation.x, sensor_state.rotation.y, sensor_state.rotation.z)
        
        quat = [sensor_state.rotation.w, sensor_state.rotation.x, sensor_state.rotation.y, sensor_state.rotation.z]
        rot = R.from_quat(quat)
        rot_euler = rot.as_euler('xyz', degrees=True)
        # print("euler: ", rot_euler)
        return sensor_state.position[0], sensor_state.position[2], rot_euler, highlighted

# save video initial
path = "video/" + object + ".mp4"
fourcc = cv2.VideoWriter_fourcc(*'mp4v')
videowriter = cv2.VideoWriter(path, fourcc, 2, (512, 512))

action = "move_forward"
x_now, y_now, theta_now, img = navigateAndSee(action)
videowriter.write(img)

nav_step = 1
nav_state = 0
while True:
    # rotate state
    if nav_state == 0:
        
        # make the angle range be -180~180 degrees, not 0~90 and 0~-90 degrees
        if theta_now[0] == 0 and theta_now[2] == 180:
            if 0<=theta_now[1]<=90:
                theta_now[1] = 180 - theta_now[1]
            elif -90<=theta_now[1]<=0:
                theta_now[1] = (-180) - theta_now[1]
        # print("euler: ", theta_now)

        # decide to turn left or right
        nav_vector = np.array(navigation_point[nav_step]) - np.array([x_now, y_now])
        theta_nav = math.atan2(nav_vector[1], nav_vector[0])/math.pi*180
        theta_face = theta_now[1]
        delta = theta_nav - theta_face + 90
        # print("delta: ", delta)

        if delta < (-180):
            delta = 360 + delta
        elif delta > 180:
            delta = -360 + delta

        if 0<=delta<180:
            print("turn right")
            action = "turn_right"
            x_now, y_now, theta_now, img = navigateAndSee(action)
        elif -180<=delta<0:
            print("turn left")
            action = "turn_left"
            x_now, y_now, theta_now, img = navigateAndSee(action)

        if -0.5<delta<0.5:
            print("turn complete")
            videowriter.write(img)
            nav_state = 1
    # go straight state
    elif nav_state == 1:
        print("go")
        action = "move_forward"
        x_now, y_now, theta_now, img = navigateAndSee(action)
        end_range = 0.1
        if navigation_point[nav_step][0]-end_range<x_now<navigation_point[nav_step][0]+end_range and navigation_point[nav_step][1]-end_range<y_now<navigation_point[nav_step][1]+end_range:
            videowriter.write(img)
            if nav_step == len(navigation_point)-1: # when arrived
                print("arrived")
                videowriter.release()
                cv2.waitKey()
                break
            else:
                print("get to one point")
                nav_state = 0
                nav_step = nav_step +1
    
    cv2.waitKey(1)